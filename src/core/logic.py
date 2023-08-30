#logic.py
# Import required modules
import openpyxl
import pymysql
import datetime
import time
import os
from tkinter import filedialog, messagebox, Toplevel
from tkinter import *
from dotenv import load_dotenv
from pathlib import Path
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from matplotlib.figure import Figure

from auth.login import username



# Specify the path to the .env file
env_path = Path('.') / 'key' / '.env'

# Load the .env file
load_dotenv(dotenv_path=env_path)


# Load variables from .env
hosts = os.environ.get('DB_HOST')
user = os.environ.get('DB_USER')
passwords = os.environ.get('DB_PASS')
bases = os.environ.get('DB_NAME')
users = os.environ.get('DB_USERS')
data = os.environ.get('DB_DATA')


def db_connect(root,table):
    global mycursor, con
    try:
        con = pymysql.connect(host=hosts, user=user, password=passwords, database=bases)
        mycursor = con.cursor()
    except Exception as e:
        messagebox.showerror('Error', f'Error connecting to database: {str(e)}', parent=root)
        return
    # Check if the database exists
    mycursor.execute("SHOW DATABASES")
    databases = [db[0] for db in mycursor.fetchall()]

    if bases not in [db.lower() for db in databases]:
        mycursor.execute(f'CREATE DATABASE {bases}')

    mycursor.execute(f'USE {bases}')

    # Check if the table exists
    mycursor.execute("SHOW TABLES")
    tables = [table[0] for table in mycursor.fetchall()]

    if data not in tables:
        query = (f"CREATE TABLE {data} ("
             "id INT AUTO_INCREMENT PRIMARY KEY, "
             "username VARCHAR(255), "
             "acct VARCHAR(255), "
             "date DATE, "
             "check_num VARCHAR(5), "
             "description VARCHAR(255), "
             "debit DECIMAL(10, 2), "
             "credit DECIMAL(10, 2), "
             "status VARCHAR(20), "
             "balance DECIMAL(15, 2), "
             "classification VARCHAR(50))")
        mycursor.execute(query)

    messagebox.showinfo('Success', 'Database Connection is successful', parent=root)
    auto_refresh(2000,root,table)
    pass

def execute_db_query(query, params=()):
    if con and mycursor:
        try:
            mycursor.execute(query, params)
            con.commit()
            return mycursor.fetchall()
        except Exception as e:
            messagebox.showerror('Error', f'Error executing query: {str(e)}')
    return None


def add_Purchase(dateEntry, descrEntry,amountEntry,classEntry,entryScreen,root):
    if not (dateEntry.get() and descrEntry.get() and amountEntry.get() and classEntry.get()):
        messagebox.showerror('Error', 'All Fields are required', parent=entryScreen)
        return
    query = (f'INSERT INTO {data} (username, date, description, debit, credit, classification, acct, check_num, status, balance) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)')
    params = (username, dateEntry.get(), descrEntry.get(), amountEntry.get(), 0, classEntry.get(), 'n/a', 'n/a', 'n/a', 0)
    
    execute_db_query(query, params)

    #Another entry
    result = messagebox.askyesno('Confirm', 'Data added successfully. Do you want to clean the form?', parent=entryScreen)
    if result:
        dateEntry.delete(0, END)
        descrEntry.delete(0, END)
        amountEntry.delete(0, END)
        classEntry.delete(0, END)

    update_table_from_database()
    pass

def update_table_from_database(root,table):
    try:
        query = (f'SELECT date, description, IFNULL(credit, 0) - IFNULL(debit, 0) as amount, balance, classification FROM {data} WHERE username = %s')
        rows = execute_db_query(query, (username,))
        if rows is not None:
            for record in table.get_children():
                table.delete(record)
            for row in rows:
                table.insert("", "end", values=row)
    except Exception as e:
        messagebox.showerror('Error', f'Error fetching data from the database: {str(e)}', parent=root)
    pass 

def auto_refresh(interval, root,table):
    update_table_from_database(root, table)
    root.after(interval, auto_refresh(interval, root, table))
    pass 


def search_Data(dateEntry, descrEntry,amountEntry,classEntry,entryScreen,table):
    try:
        #Seacrch based on 
        query = (f'SELECT date, description, IFNULL(debit, 0) - IFNULL(credit, 0) as amount, balance, classification FROM {data} WHERE username = %s')
        params = [username]

        #check which thing
        if dateEntry.get():
            query += " AND date = %s"
            params.append(dateEntry.get())
        if descrEntry.get():
            query += " AND description LIKE %s"
            params.append(f"%{descrEntry.get()}%")
        if amountEntry.get():
            query += " AND (credit = %s OR debit = %s)"
            params.extend([amountEntry.get(), amountEntry.get()])
        if classEntry.get():
            query += " AND classification = %s"
            params.append(classEntry.get())

        # Execute the query
        rows = execute_db_query(query, tuple(params))
        
        # Clear current contents of the table
        for record in table.get_children():
            table.delete(record)

        # Insert the new records into the table
        if rows is not None:
            for row in rows:
                table.insert("", "end", values=row)

    except Exception as e:
        messagebox.showerror('Error', f'Error fetching data from the database: {str(e)}', parent=entryScreen)
    pass 
def remove_Purchase(table , root):
    try:
        selected_items = table.selection() 
        for selected_item in selected_items:
            values = table.item(selected_item, 'values')


            if len(values) < 2:
                continue

            # Extract values from the selected row
            selected_date, selected_description, _, _, _ = values

            # Check if records exist before deleting
            query = (f'SELECT * FROM {data} WHERE username = %s AND date = %s AND description = %s')
            params = (username, selected_date, selected_description)
            rows = execute_db_query(query, params)

            if not rows:
                #messagebox.showinfo('Info', 'No records found to delete.', parent=root)
                continue

            # Delete the record
            query = (f'DELETE FROM {data} WHERE username = %s AND date = %s AND description = %s')
            params = (username, selected_date, selected_description)
            execute_db_query(query, params)

        # Update the table to reflect the changes
        update_table_from_database()

        messagebox.showinfo('Success', 'Record deleted successfully.', parent=root)

    except Exception as e:
        messagebox.showerror('Error', f'Error deleting record: {str(e)}', parent=root)
    pass 

def display_Data(root):
    try:

        # Fetch time-series data for line chart
        query = (f'SELECT date, IFNULL(credit, 0) - IFNULL(debit, 0) as amount FROM {data} WHERE username = %s ORDER BY date')
        mycursor.execute(query, (username,))
        rows = mycursor.fetchall()
        dates = [row[0] for row in rows]
        amounts = [row[1] for row in rows]

        # Fetch data for bar chart and pie chart
        query = (f'SELECT classification, SUM(IFNULL(credit, 0) - IFNULL(debit, 0)) as amount FROM {data} WHERE username = %s GROUP BY classification')
        mycursor.execute(query, (username,))
        categories = mycursor.fetchall()

        #Filter out categories with negative or zero amounts
        filtered_categories = [(label, value) for label, value in categories if value > 0]
        if not filtered_categories:
            raise ValueError("No positive values to plot.")

        labels = [category[0] for category in filtered_categories]
        values = [category[1] for category in filtered_categories]


        displayWindow = Toplevel(root)
        displayWindow.title('Spending Display')
        
        # Create matplotlib Figure and three subplots
        fig = Figure(figsize=(15, 5), dpi=100)
        ax1 = fig.add_subplot(131)
        ax2 = fig.add_subplot(132)
        ax3 = fig.add_subplot(133)

        # Line Chart
        ax1.plot(dates, amounts)
        ax1.set_title('Amount Spent Over Time', fontsize=12)
        ax1.set_xlabel('Date', fontsize=10)
        ax1.set_ylabel('Amount', fontsize=10)
        ax1.tick_params(axis='both', which='major', labelsize=8)
        for tick in ax1.get_xticklabels():
            tick.set_rotation(90)  # Rotate x-axis text labels for Line Chart

        # Bar Chart
        ax2.bar(labels, values)
        ax2.set_title('Spending Per Category', fontsize=12)
        ax2.set_xlabel('Category', fontsize=10)
        ax2.set_ylabel('Amount', fontsize=10)
        ax2.tick_params(axis='both', which='major', labelsize=8)
        for tick in ax2.get_xticklabels():
            tick.set_rotation(90)  # Rotate x-axis text labels for Bar Chart

        # Pie Chart
        if any(v <= 0 for v in values):
            raise ValueError("All values must be positive for a pie chart.")
        ax3.pie(values, labels=labels, autopct='%1.1f%%')
        ax3.set_title('Spending by Category', fontsize=12)

        # Embed the matplotlib Figure in the Tkinter window
        canvas = FigureCanvasTkAgg(fig, master=displayWindow)  # A tk.DrawingArea.
        canvas.draw()
        canvas.get_tk_widget().pack(side='top', fill='both', expand=1)


    except Exception as e:
        messagebox.showerror('Error', f'Error fetching data for display: {str(e)}', parent=root)

    pass 
def load_Data(root):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")])
    if file_path:
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                
                acct_number = row[0]  or "Unknown"
                # Convert date format
                if isinstance(row[1], datetime.datetime):
                    date = row[1]
                elif row[1]:
                    date = datetime.datetime.strptime(row[1], '%m/%d/%Y').date()
                else:
                    date = datetime.date.today()
                check_value = row[2] or None
                description = row[3] or "Unknown"
                debit = row[4] or 0
                credit = row[5] or 0
                status = row[6] or "Unknown"
                balance = row[7] or 0
                classification = row[8] or "Uncategorized"

                # Check duplicates
                check_query = f"SELECT * FROM {data} WHERE username = %s AND date = %s AND description = %s"
                check_values = (username, date, description)
                mycursor.execute(check_query, check_values)
                existing_record = mycursor.fetchone()

                if existing_record:
                    continue  # Skip this record if it already exists
                query = (f'INSERT INTO {data} (username, acct, date, check_num, description, debit, credit, status, balance, classification) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)')
                values = (username, acct_number, date, check_value, description, debit, credit, status, balance, classification)
                mycursor.execute(query, values)

            con.commit()

            messagebox.showinfo('Success', 'Data loaded from Excel file and inserted into the database', parent=root)

        except Exception as e:
            messagebox.showerror('Error', f'Error loading data and inserting into the database: {str(e)}', parent=root)
        pass 


def export_Data(table, root):
    saveas= filedialog.asksaveasfilename(defaultextension='.xlsx')
    if not saveas:# if user cancelled
        return
    tab=table.get_children()
    savelist=[]

    #store the data from the table
    for i in tab:
        data=table.item(i)
        datas=data['values']
        savelist.append(datas)

    #worksheet to add into 
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write first row
    headers = ['Date', 'Description', 'Amount', 'Balance', 'Classification']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write all the data from the table into the sheets
    for row_num, row_data in enumerate(savelist, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    # Save to the specified file
    workbook.save(saveas)
    messagebox.showinfo('Success', 'Data exported successfully!', parent=root)

    pass 

def Exit(root):
    result=messagebox.askyesno('Confirm','Do you want to exit?')
    if result:
        root.destroy()
        mycursor.close()
        con.close()
    else:
        pass



